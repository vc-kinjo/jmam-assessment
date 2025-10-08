from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
from datetime import datetime, date, timedelta
import io
import base64
from pathlib import Path

# PDF生成用
from weasyprint import HTML, CSS
from jinja2 import Template

# Excel生成用  
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference

from ..models.project import Project
from ..models.task import Task, TaskAssignment
from ..models.user import User

class ReportService:
    def __init__(self):
        self.template_dir = Path(__file__).parent.parent / "templates" / "reports"
        self.template_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_project_progress_report(self, db: Session, project_id: int, 
                                       report_date: Optional[date] = None) -> Dict[str, Any]:
        """プロジェクト進捗レポートを生成"""
        if report_date is None:
            report_date = datetime.now().date()
        
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            raise ValueError("Project not found")
        
        tasks = db.query(Task).filter(Task.project_id == project_id).all()
        
        # 基本統計
        total_tasks = len(tasks)
        completed_tasks = len([t for t in tasks if t.status == "completed"])
        in_progress_tasks = len([t for t in tasks if t.status == "in_progress"])
        not_started_tasks = len([t for t in tasks if t.status == "not_started"])
        overdue_tasks = len([t for t in tasks 
                           if t.planned_end_date and t.planned_end_date < report_date 
                           and t.status != "completed"])
        
        # 進捗率計算
        completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
        
        # 工数統計
        total_estimated_hours = sum(t.estimated_hours or 0 for t in tasks)
        total_actual_hours = sum(t.actual_hours or 0 for t in tasks)
        
        # マイルストーン
        milestones = [t for t in tasks if t.is_milestone]
        completed_milestones = [m for m in milestones if m.status == "completed"]
        
        # 遅延分析
        delayed_tasks = []
        for task in tasks:
            if (task.planned_end_date and task.planned_end_date < report_date 
                and task.status != "completed"):
                delay_days = (report_date - task.planned_end_date).days
                delayed_tasks.append({
                    "name": task.name,
                    "planned_end_date": task.planned_end_date,
                    "delay_days": delay_days,
                    "status": task.status
                })
        
        # カテゴリ別進捗
        category_progress = {}
        for task in tasks:
            category = task.category or "その他"
            if category not in category_progress:
                category_progress[category] = {"total": 0, "completed": 0}
            
            category_progress[category]["total"] += 1
            if task.status == "completed":
                category_progress[category]["completed"] += 1
        
        # 進捗率を計算
        for category in category_progress:
            total = category_progress[category]["total"]
            completed = category_progress[category]["completed"]
            category_progress[category]["completion_rate"] = (completed / total * 100) if total > 0 else 0
        
        return {
            "project": {
                "id": project.id,
                "name": project.name,
                "start_date": project.start_date,
                "end_date": project.end_date,
                "status": project.status
            },
            "report_date": report_date,
            "summary": {
                "total_tasks": total_tasks,
                "completed_tasks": completed_tasks,
                "in_progress_tasks": in_progress_tasks,
                "not_started_tasks": not_started_tasks,
                "overdue_tasks": overdue_tasks,
                "completion_rate": round(completion_rate, 1),
                "total_estimated_hours": total_estimated_hours,
                "total_actual_hours": total_actual_hours,
                "milestones_total": len(milestones),
                "milestones_completed": len(completed_milestones)
            },
            "delayed_tasks": delayed_tasks,
            "category_progress": category_progress
        }
    
    def export_project_progress_to_pdf(self, db: Session, project_id: int, 
                                     report_date: Optional[date] = None) -> bytes:
        """プロジェクト進捗レポートをPDF出力"""
        report_data = self.generate_project_progress_report(db, project_id, report_date)
        
        # HTMLテンプレート
        template_html = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>プロジェクト進捗レポート</title>
            <style>
                body { font-family: 'Noto Sans JP', sans-serif; margin: 20px; }
                .header { text-align: center; border-bottom: 2px solid #333; padding-bottom: 10px; }
                .section { margin: 20px 0; }
                .summary-table { width: 100%; border-collapse: collapse; margin: 10px 0; }
                .summary-table th, .summary-table td { 
                    border: 1px solid #ddd; padding: 8px; text-align: left; 
                }
                .summary-table th { background-color: #f2f2f2; }
                .progress-bar { 
                    width: 200px; height: 20px; background-color: #f0f0f0; 
                    border: 1px solid #ccc; position: relative; 
                }
                .progress-fill { 
                    height: 100%; background-color: #4CAF50; 
                }
                .delayed-task { background-color: #ffebee; }
                .chart-placeholder { 
                    width: 100%; height: 200px; border: 1px solid #ccc; 
                    text-align: center; line-height: 200px; background-color: #f9f9f9; 
                }
            </style>
        </head>
        <body>
            <div class="header">
                <h1>プロジェクト進捗レポート</h1>
                <h2>{{ project.name }}</h2>
                <p>レポート作成日: {{ report_date.strftime('%Y年%m月%d日') }}</p>
            </div>
            
            <div class="section">
                <h3>プロジェクト概要</h3>
                <table class="summary-table">
                    <tr><td>プロジェクト名</td><td>{{ project.name }}</td></tr>
                    <tr><td>開始日</td><td>{{ project.start_date.strftime('%Y/%m/%d') if project.start_date else '未設定' }}</td></tr>
                    <tr><td>終了予定日</td><td>{{ project.end_date.strftime('%Y/%m/%d') if project.end_date else '未設定' }}</td></tr>
                    <tr><td>ステータス</td><td>{{ project.status }}</td></tr>
                </table>
            </div>
            
            <div class="section">
                <h3>進捗サマリー</h3>
                <table class="summary-table">
                    <tr><td>総タスク数</td><td>{{ summary.total_tasks }}</td></tr>
                    <tr><td>完了タスク</td><td>{{ summary.completed_tasks }}</td></tr>
                    <tr><td>進行中タスク</td><td>{{ summary.in_progress_tasks }}</td></tr>
                    <tr><td>未開始タスク</td><td>{{ summary.not_started_tasks }}</td></tr>
                    <tr><td>遅延タスク</td><td>{{ summary.overdue_tasks }}</td></tr>
                    <tr>
                        <td>完了率</td>
                        <td>
                            {{ summary.completion_rate }}%
                            <div class="progress-bar">
                                <div class="progress-fill" style="width: {{ summary.completion_rate }}%;"></div>
                            </div>
                        </td>
                    </tr>
                    <tr><td>予定工数合計</td><td>{{ summary.total_estimated_hours }}時間</td></tr>
                    <tr><td>実績工数合計</td><td>{{ summary.total_actual_hours }}時間</td></tr>
                </table>
            </div>
            
            {% if delayed_tasks %}
            <div class="section">
                <h3>遅延タスク一覧</h3>
                <table class="summary-table">
                    <thead>
                        <tr>
                            <th>タスク名</th>
                            <th>予定終了日</th>
                            <th>遅延日数</th>
                            <th>ステータス</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for task in delayed_tasks %}
                        <tr class="delayed-task">
                            <td>{{ task.name }}</td>
                            <td>{{ task.planned_end_date.strftime('%Y/%m/%d') if task.planned_end_date else '-' }}</td>
                            <td>{{ task.delay_days }}日</td>
                            <td>{{ task.status }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
            {% endif %}
            
            <div class="section">
                <h3>カテゴリ別進捗</h3>
                <table class="summary-table">
                    <thead>
                        <tr>
                            <th>カテゴリ</th>
                            <th>総タスク数</th>
                            <th>完了タスク数</th>
                            <th>完了率</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for category, progress in category_progress.items() %}
                        <tr>
                            <td>{{ category }}</td>
                            <td>{{ progress.total }}</td>
                            <td>{{ progress.completed }}</td>
                            <td>{{ progress.completion_rate|round(1) }}%</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        </body>
        </html>
        """
        
        # テンプレートを使ってHTMLを生成
        template = Template(template_html)
        html_content = template.render(**report_data)
        
        # PDFに変換
        pdf_bytes = HTML(string=html_content).write_pdf()
        return pdf_bytes
    
    def export_tasks_to_excel(self, db: Session, project_id: int) -> bytes:
        """タスク一覧をExcelエクスポート"""
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            raise ValueError("Project not found")
        
        tasks = db.query(Task).filter(Task.project_id == project_id).order_by(Task.sort_order).all()
        
        # Excelワークブック作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "タスク一覧"
        
        # ヘッダー設定
        headers = [
            "ID", "タスク名", "説明", "ステータス", "優先度", "カテゴリ",
            "予定開始日", "予定終了日", "実績開始日", "実績終了日",
            "予定工数(h)", "実績工数(h)", "進捗率(%)", "マイルストーン", "親タスク"
        ]
        
        # ヘッダーのスタイル
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"), 
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # ヘッダー行を設定
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # データ行を設定
        for row, task in enumerate(tasks, 2):
            # 親タスク名を取得
            parent_task_name = ""
            if task.parent_task_id:
                parent_task = db.query(Task).filter(Task.id == task.parent_task_id).first()
                parent_task_name = parent_task.name if parent_task else ""
            
            data = [
                task.id,
                task.name,
                task.description or "",
                task.status,
                task.priority,
                task.category or "",
                task.planned_start_date.strftime('%Y/%m/%d') if task.planned_start_date else "",
                task.planned_end_date.strftime('%Y/%m/%d') if task.planned_end_date else "",
                task.actual_start_date.strftime('%Y/%m/%d') if task.actual_start_date else "",
                task.actual_end_date.strftime('%Y/%m/%d') if task.actual_end_date else "",
                task.estimated_hours or 0,
                task.actual_hours or 0,
                task.progress_rate,
                "はい" if task.is_milestone else "いいえ",
                parent_task_name
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                
                # ステータスに応じて背景色を設定
                if col == 4:  # ステータス列
                    if value == "completed":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "in_progress":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == "overdue":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 列幅を自動調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # プロジェクト情報シートを追加
        project_ws = wb.create_sheet(title="プロジェクト情報")
        project_info = [
            ["項目", "値"],
            ["プロジェクト名", project.name],
            ["説明", project.description or ""],
            ["開始日", project.start_date.strftime('%Y/%m/%d') if project.start_date else ""],
            ["終了予定日", project.end_date.strftime('%Y/%m/%d') if project.end_date else ""],
            ["ステータス", project.status],
            ["カテゴリ", project.category or ""],
            ["作成日", project.created_at.strftime('%Y/%m/%d %H:%M') if project.created_at else ""],
        ]
        
        for row, (key, value) in enumerate(project_info, 1):
            project_ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            project_ws.cell(row=row, column=2, value=value)
        
        # Excelファイルをバイト列として返す
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    def generate_user_workload_report(self, db: Session, start_date: date, end_date: date) -> Dict[str, Any]:
        """ユーザー別作業負荷レポートを生成"""
        users = db.query(User).all()
        
        workload_data = []
        
        for user in users:
            # 期間内のタスク割り当てを取得
            assignments = db.query(TaskAssignment).join(Task).filter(
                TaskAssignment.user_id == user.id,
                Task.planned_start_date <= end_date,
                Task.planned_end_date >= start_date
            ).all()
            
            total_estimated_hours = sum(assignment.task.estimated_hours or 0 for assignment in assignments)
            total_actual_hours = sum(assignment.task.actual_hours or 0 for assignment in assignments)
            
            active_tasks = len([a for a in assignments if a.task.status in ["not_started", "in_progress"]])
            completed_tasks = len([a for a in assignments if a.task.status == "completed"])
            overdue_tasks = len([a for a in assignments 
                               if a.task.planned_end_date and a.task.planned_end_date < datetime.now().date() 
                               and a.task.status != "completed"])
            
            workload_data.append({
                "user_id": user.id,
                "user_name": user.full_name,
                "email": user.email,
                "role": user.role_level,
                "total_tasks": len(assignments),
                "active_tasks": active_tasks,
                "completed_tasks": completed_tasks,
                "overdue_tasks": overdue_tasks,
                "total_estimated_hours": total_estimated_hours,
                "total_actual_hours": total_actual_hours,
                "efficiency_rate": (total_actual_hours / total_estimated_hours * 100) 
                                 if total_estimated_hours > 0 else 0
            })
        
        return {
            "period": {
                "start_date": start_date,
                "end_date": end_date
            },
            "workload_data": workload_data
        }

report_service = ReportService()